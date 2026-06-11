from __future__ import annotations

import base64
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st

from core.models import (
    Analysis,
    Alternative,
    Criterion,
    Level,
    closest_level_code,
    level_score,
    neutral_level_code,
    scale_for_analysis,
)
from core.storage import DATA_DIR, analysis_to_json, list_analysis_files, load_analysis, load_analysis_from_json_text, save_analysis
from core.sample_data import create_fruit_supplier_analysis
from core.calculation import criterion_contributions, evaluate_all, evaluate_phase, normalize_weights, robustness_summary, selected_criteria, sensitivity_scenarios


DARK_BLUE = "#173A5E"
TEAL = "#178F88"
BRIGHT_TEAL = "#2AC6B6"
GREEN = "#49B051"
LIGHT_BG = "#F3F8F4"
MENU_ITEMS = [
    "Dashboard",
    "Análise",
    "Fornecedores",
    "Critérios & Pesos",
    "Escala",
    "Avaliações",
    "Resultados",
    "Sensibilidade",
]
MENU_KEYS = [
    "nav_dashboard",
    "nav_analysis",
    "nav_suppliers",
    "nav_criteria",
    "nav_levels",
    "nav_scores",
    "nav_results",
    "nav_sensitivity",
]
LANGUAGES = {
    "pt": "Português",
    "en": "English",
    "es": "Español",
    "tr": "Türkçe",
}
TRANSLATIONS = {
    "pt": {
        "current_analysis": "Análise atual",
        "language": "Idioma",
        "new_analysis": "Nova Análise",
        "read_analysis": "Ler Análise",
        "save_analysis": "Gravar Análise",
        "producer_filter": "Produtores nos gráficos",
        "producer_filter_help": "Limita apenas a visualização. Os cálculos continuam a usar todos os produtores.",
        "all": "Todos",
        "complete_export": "Exportação completa",
        "export_csv": "Exportar CSV",
        "export_excel": "Exportar Excel",
        "export_pdf": "Exportar PDF",
        "yes": "Sim",
        "no": "Não",
        "cancel": "Cancelar",
        "create": "Criar",
        "save": "Gravar",
        "saved_to": "Guardado em {name}",
        "new_analysis_created": "Nova análise criada.",
        "analysis_loaded": "Análise carregada.",
        "confirm_new_body": "Antes de criar uma nova análise, escolha o que fazer com a análise atual.",
        "confirm_read_body": "Antes de ler outra análise JSON, escolha o que fazer com a análise atual.",
        "new_analysis_name": "Nome da nova análise",
        "save_filename": "Nome do ficheiro JSON",
        "select_json": "Selecionar ficheiro JSON",
        "select_json_error": "Selecione um ficheiro JSON.",
        "read_error": "Erro ao ler análise: {error}",
        "nav_dashboard": "Dashboard",
        "nav_analysis": "Análise",
        "nav_suppliers": "Fornecedores",
        "nav_criteria": "Critérios & Pesos",
        "nav_levels": "Escala",
        "nav_scores": "Avaliações",
        "nav_results": "Resultados",
        "nav_sensitivity": "Sensibilidade",
        "problem": "Problema",
        "decision_conflicts": "Conflitos de decisão",
        "decision_conflicts_text": "Preço vs. qualidade, prazo vs. variedade, sustentabilidade vs. custo e reputação vs. proximidade.",
        "scale": "Escala",
        "scale_text": "Escala semântica configurável, com códigos, pontuações e cores definidos pela análise.",
        "quick_nav": "Navegação rápida e fluxo principal na barra superior.",
        "steps_title": "8 etapas MCDA",
        "recommended_supplier": "Fornecedor recomendado",
        "global_value": "Valor Global",
        "global_value_axis": "Valor global",
        "supplier": "Fornecedor",
        "final_ranking": "Ranking final",
        "analysis_name": "Nome da análise",
        "description": "Descrição",
        "context": "Contexto do problema",
        "notes": "Notas",
        "methodology": "Metodologia",
        "owner": "Responsável",
        "alternatives": "Alternativas",
        "criteria": "Critérios",
        "save_analysis_changes": "Guardar alterações da análise",
        "cost": "Custo",
        "apply_suppliers": "Aplicar fornecedores",
        "criteria_help": "Adicione linhas para criar critérios; edite linhas existentes; apague linhas para remover critérios.",
        "code": "Código",
        "name": "Nome",
        "phase": "Fase",
        "selected": "Selecionado",
        "weight": "Peso",
        "cost_risk": "Custo/Risco?",
        "order": "Ordem",
        "unique_criterion_help": "Identificador único do critério, por exemplo A1.7.",
        "apply_criteria": "Aplicar critérios e pesos",
        "original_weight": "Peso original",
        "normalized_weight_pct": "Peso normalizado (%)",
        "relative_importance": "Importância relativa dos critérios",
        "level": "Nível",
        "score": "Pontuação",
        "semantic_scale_title": "Escala semântica",
        "reference_example": "Exemplo de referência: Calibre",
        "neutral": "Neutro",
        "better": "Melhor",
        "worse": "Pior",
        "apply_scores": "Aplicar avaliações",
        "colored_matrix": "Matriz de avaliação colorida",
        "matrix_caption": "Produtores nas linhas e critérios nas colunas. A lista segue o filtro visual definido na sidebar.",
        "producer": "Produtor",
        "no_results": "Não existem critérios selecionados para calcular resultados.",
        "formula": "Fórmula",
        "formula_text": "Valor Global = Σ (Peso normalizado do Critério × Pontuação na Escala)",
        "contribution": "Contribuição",
        "criterion": "Critério",
        "criteria_contribution": "Contribuição dos critérios por fornecedor",
        "download_phase_csv": "Descarregar resultados da fase {phase} em CSV",
        "no_sensitivity": "Não existem dados suficientes para análise de sensibilidade.",
        "scenarios": "Cenários",
        "scenario": "Cenário",
        "ranking_stability": "Estabilidade do ranking por cenário",
        "robustness_summary": "Resumo de robustez",
        "keeps_base_winner": "Mantém vencedor base?",
        "winner": "Vencedor",
        "robust_conclusion": "Conclusão: ranking robusto — {winner} mantém a recomendação nos cenários testados.",
        "sensitive_ranking": "O ranking é sensível em pelo menos um cenário.",
        "selected_phase2": "Selecionada para 2.ª fase",
    },
    "en": {
        "current_analysis": "Current analysis", "language": "Language", "new_analysis": "New Analysis", "read_analysis": "Open Analysis", "save_analysis": "Save Analysis", "producer_filter": "Producers in charts", "producer_filter_help": "Only limits the view. Calculations still use all producers.", "all": "All", "complete_export": "Complete export", "export_csv": "Export CSV", "export_excel": "Export Excel", "export_pdf": "Export PDF", "yes": "Yes", "no": "No", "cancel": "Cancel", "create": "Create", "save": "Save", "saved_to": "Saved to {name}", "new_analysis_created": "New analysis created.", "analysis_loaded": "Analysis loaded.", "confirm_new_body": "Before creating a new analysis, choose what to do with the current analysis.", "confirm_read_body": "Before opening another JSON analysis, choose what to do with the current analysis.", "new_analysis_name": "New analysis name", "save_filename": "JSON file name", "select_json": "Select JSON file", "select_json_error": "Select a JSON file.", "read_error": "Error reading analysis: {error}", "nav_dashboard": "Dashboard", "nav_analysis": "Analysis", "nav_suppliers": "Suppliers", "nav_criteria": "Criteria & Weights", "nav_levels": "Scale", "nav_scores": "Assessments", "nav_results": "Results", "nav_sensitivity": "Sensitivity", "problem": "Problem", "decision_conflicts": "Decision conflicts", "decision_conflicts_text": "Price vs. quality, lead time vs. variety, sustainability vs. cost, and reputation vs. proximity.", "scale": "Scale", "scale_text": "Configurable semantic scale with codes, scores, and colors defined by the analysis.", "quick_nav": "Quick navigation and main workflow in the top bar.", "steps_title": "8 MCDA steps", "recommended_supplier": "Recommended supplier", "global_value": "Global Value", "global_value_axis": "Global value", "supplier": "Supplier", "final_ranking": "Final ranking", "analysis_name": "Analysis name", "description": "Description", "context": "Problem context", "notes": "Notes", "methodology": "Methodology", "owner": "Owner", "alternatives": "Alternatives", "criteria": "Criteria", "save_analysis_changes": "Save analysis changes", "cost": "Cost", "apply_suppliers": "Apply suppliers", "criteria_help": "Add rows to create criteria; edit existing rows; delete rows to remove criteria.", "code": "Code", "name": "Name", "phase": "Phase", "selected": "Selected", "weight": "Weight", "cost_risk": "Cost/Risk?", "order": "Order", "unique_criterion_help": "Unique criterion identifier, for example A1.7.", "apply_criteria": "Apply criteria and weights", "original_weight": "Original weight", "normalized_weight_pct": "Normalized weight (%)", "relative_importance": "Relative importance of criteria", "level": "Level", "score": "Score", "semantic_scale_title": "Semantic scale", "reference_example": "Reference example: Size grade", "neutral": "Neutral", "better": "Better", "worse": "Worse", "apply_scores": "Apply assessments", "colored_matrix": "Colored assessment matrix", "matrix_caption": "Producers in rows and criteria in columns. The list follows the visual filter set in the sidebar.", "producer": "Producer", "no_results": "No selected criteria are available to calculate results.", "formula": "Formula", "formula_text": "Global Value = Σ (Normalized Criterion Weight × Scale Score)", "contribution": "Contribution", "criterion": "Criterion", "criteria_contribution": "Criteria contribution by supplier", "download_phase_csv": "Download phase {phase} results as CSV", "no_sensitivity": "There is not enough data for sensitivity analysis.", "scenarios": "Scenarios", "scenario": "Scenario", "ranking_stability": "Ranking stability by scenario", "robustness_summary": "Robustness summary", "keeps_base_winner": "Keeps base winner?", "winner": "Winner", "robust_conclusion": "Conclusion: robust ranking — {winner} keeps the recommendation in the tested scenarios.", "sensitive_ranking": "The ranking is sensitive in at least one scenario.", "selected_phase2": "Selected for phase 2",
    },
    "es": {
        "current_analysis": "Análisis actual", "language": "Idioma", "new_analysis": "Nuevo Análisis", "read_analysis": "Leer Análisis", "save_analysis": "Guardar Análisis", "producer_filter": "Productores en gráficos", "producer_filter_help": "Solo limita la visualización. Los cálculos siguen usando todos los productores.", "all": "Todos", "complete_export": "Exportación completa", "export_csv": "Exportar CSV", "export_excel": "Exportar Excel", "export_pdf": "Exportar PDF", "yes": "Sí", "no": "No", "cancel": "Cancelar", "create": "Crear", "save": "Guardar", "saved_to": "Guardado en {name}", "new_analysis_created": "Nuevo análisis creado.", "analysis_loaded": "Análisis cargado.", "confirm_new_body": "Antes de crear un nuevo análisis, elija qué hacer con el análisis actual.", "confirm_read_body": "Antes de leer otro análisis JSON, elija qué hacer con el análisis actual.", "new_analysis_name": "Nombre del nuevo análisis", "save_filename": "Nombre del fichero JSON", "select_json": "Seleccionar fichero JSON", "select_json_error": "Seleccione un fichero JSON.", "read_error": "Error al leer el análisis: {error}", "nav_dashboard": "Panel", "nav_analysis": "Análisis", "nav_suppliers": "Proveedores", "nav_criteria": "Criterios y Pesos", "nav_levels": "Escala", "nav_scores": "Evaluaciones", "nav_results": "Resultados", "nav_sensitivity": "Sensibilidad", "problem": "Problema", "decision_conflicts": "Conflictos de decisión", "decision_conflicts_text": "Precio vs. calidad, plazo vs. variedad, sostenibilidad vs. coste y reputación vs. proximidad.", "scale": "Escala", "scale_text": "Escala semántica configurable con códigos, puntuaciones y colores definidos por el análisis.", "quick_nav": "Navegación rápida y flujo principal en la barra superior.", "steps_title": "8 etapas MCDA", "recommended_supplier": "Proveedor recomendado", "global_value": "Valor Global", "global_value_axis": "Valor global", "supplier": "Proveedor", "final_ranking": "Ranking final", "analysis_name": "Nombre del análisis", "description": "Descripción", "context": "Contexto del problema", "notes": "Notas", "methodology": "Metodología", "owner": "Responsable", "alternatives": "Alternativas", "criteria": "Criterios", "save_analysis_changes": "Guardar cambios del análisis", "cost": "Coste", "apply_suppliers": "Aplicar proveedores", "criteria_help": "Añada filas para crear criterios; edite filas existentes; elimine filas para quitar criterios.", "code": "Código", "name": "Nombre", "phase": "Fase", "selected": "Seleccionado", "weight": "Peso", "cost_risk": "Coste/Riesgo?", "order": "Orden", "unique_criterion_help": "Identificador único del criterio, por ejemplo A1.7.", "apply_criteria": "Aplicar criterios y pesos", "original_weight": "Peso original", "normalized_weight_pct": "Peso normalizado (%)", "relative_importance": "Importancia relativa de los criterios", "level": "Nivel", "score": "Puntuación", "semantic_scale_title": "Escala semántica", "reference_example": "Ejemplo de referencia: Calibre", "neutral": "Neutro", "better": "Mejor", "worse": "Peor", "apply_scores": "Aplicar evaluaciones", "colored_matrix": "Matriz de evaluación coloreada", "matrix_caption": "Productores en las filas y criterios en las columnas. La lista sigue el filtro visual definido en la barra lateral.", "producer": "Productor", "no_results": "No hay criterios seleccionados para calcular resultados.", "formula": "Fórmula", "formula_text": "Valor Global = Σ (Peso normalizado del Criterio × Puntuación en la Escala)", "contribution": "Contribución", "criterion": "Criterio", "criteria_contribution": "Contribución de criterios por proveedor", "download_phase_csv": "Descargar resultados de la fase {phase} en CSV", "no_sensitivity": "No hay datos suficientes para el análisis de sensibilidad.", "scenarios": "Escenarios", "scenario": "Escenario", "ranking_stability": "Estabilidad del ranking por escenario", "robustness_summary": "Resumen de robustez", "keeps_base_winner": "Mantiene el ganador base?", "winner": "Ganador", "robust_conclusion": "Conclusión: ranking robusto — {winner} mantiene la recomendación en los escenarios probados.", "sensitive_ranking": "El ranking es sensible en al menos un escenario.", "selected_phase2": "Seleccionada para la fase 2",
    },
    "tr": {
        "current_analysis": "Geçerli analiz", "language": "Dil", "new_analysis": "Yeni Analiz", "read_analysis": "Analiz Aç", "save_analysis": "Analizi Kaydet", "producer_filter": "Grafiklerdeki üreticiler", "producer_filter_help": "Yalnızca görünümü sınırlar. Hesaplamalar tüm üreticileri kullanmaya devam eder.", "all": "Tümü", "complete_export": "Tam dışa aktarma", "export_csv": "CSV dışa aktar", "export_excel": "Excel dışa aktar", "export_pdf": "PDF dışa aktar", "yes": "Evet", "no": "Hayır", "cancel": "İptal", "create": "Oluştur", "save": "Kaydet", "saved_to": "{name} dosyasına kaydedildi", "new_analysis_created": "Yeni analiz oluşturuldu.", "analysis_loaded": "Analiz yüklendi.", "confirm_new_body": "Yeni analiz oluşturmadan önce geçerli analizle ne yapılacağını seçin.", "confirm_read_body": "Başka bir JSON analizi açmadan önce geçerli analizle ne yapılacağını seçin.", "new_analysis_name": "Yeni analiz adı", "save_filename": "JSON dosya adı", "select_json": "JSON dosyası seç", "select_json_error": "Bir JSON dosyası seçin.", "read_error": "Analiz okunurken hata: {error}", "nav_dashboard": "Panel", "nav_analysis": "Analiz", "nav_suppliers": "Tedarikçiler", "nav_criteria": "Kriterler ve Ağırlıklar", "nav_levels": "Ölçek", "nav_scores": "Değerlendirmeler", "nav_results": "Sonuçlar", "nav_sensitivity": "Duyarlılık", "problem": "Problem", "decision_conflicts": "Karar çatışmaları", "decision_conflicts_text": "Fiyat ve kalite, teslim süresi ve çeşitlilik, sürdürülebilirlik ve maliyet, itibar ve yakınlık.", "scale": "Ölçek", "scale_text": "Kodları, puanları ve renkleri analiz tarafından tanımlanan yapılandırılabilir anlamsal ölçek.", "quick_nav": "Üst çubukta hızlı gezinme ve ana akış.", "steps_title": "8 MCDA adımı", "recommended_supplier": "Önerilen tedarikçi", "global_value": "Global Değer", "global_value_axis": "Global değer", "supplier": "Tedarikçi", "final_ranking": "Nihai sıralama", "analysis_name": "Analiz adı", "description": "Açıklama", "context": "Problem bağlamı", "notes": "Notlar", "methodology": "Metodoloji", "owner": "Sorumlu", "alternatives": "Alternatifler", "criteria": "Kriterler", "save_analysis_changes": "Analiz değişikliklerini kaydet", "cost": "Maliyet", "apply_suppliers": "Tedarikçileri uygula", "criteria_help": "Kriter oluşturmak için satır ekleyin; mevcut satırları düzenleyin; kriter silmek için satırları kaldırın.", "code": "Kod", "name": "Ad", "phase": "Aşama", "selected": "Seçili", "weight": "Ağırlık", "cost_risk": "Maliyet/Risk?", "order": "Sıra", "unique_criterion_help": "Benzersiz kriter tanımlayıcısı, örneğin A1.7.", "apply_criteria": "Kriterleri ve ağırlıkları uygula", "original_weight": "Orijinal ağırlık", "normalized_weight_pct": "Normalize ağırlık (%)", "relative_importance": "Kriterlerin göreli önemi", "level": "Seviye", "score": "Puan", "semantic_scale_title": "Anlamsal ölçek", "reference_example": "Referans örnek: Kalibre", "neutral": "Nötr", "better": "Daha iyi", "worse": "Daha kötü", "apply_scores": "Değerlendirmeleri uygula", "colored_matrix": "Renkli değerlendirme matrisi", "matrix_caption": "Üreticiler satırlarda, kriterler sütunlarda gösterilir. Liste yan çubuktaki görsel filtreyi izler.", "producer": "Üretici", "no_results": "Sonuç hesaplamak için seçili kriter yok.", "formula": "Formül", "formula_text": "Global Değer = Σ (Normalize Kriter Ağırlığı × Ölçek Puanı)", "contribution": "Katkı", "criterion": "Kriter", "criteria_contribution": "Tedarikçiye göre kriter katkısı", "download_phase_csv": "{phase}. aşama sonuçlarını CSV indir", "no_sensitivity": "Duyarlılık analizi için yeterli veri yok.", "scenarios": "Senaryolar", "scenario": "Senaryo", "ranking_stability": "Senaryoya göre sıralama kararlılığı", "robustness_summary": "Sağlamlık özeti", "keeps_base_winner": "Temel kazanan korunuyor mu?", "winner": "Kazanan", "robust_conclusion": "Sonuç: sağlam sıralama — {winner} test edilen senaryolarda öneriyi koruyor.", "sensitive_ranking": "Sıralama en az bir senaryoda duyarlıdır.", "selected_phase2": "2. aşama için seçili",
    },
}


def current_language() -> str:
    return st.session_state.get("language", "pt")


def t(key: str, **kwargs) -> str:
    text = TRANSLATIONS.get(current_language(), TRANSLATIONS["pt"]).get(key, TRANSLATIONS["pt"].get(key, key))
    return text.format(**kwargs) if kwargs else text


def translated_menu_items() -> list[str]:
    return [t(key) for key in MENU_KEYS]


def level_name_for_ui(level_code: str, fallback: str) -> str:
    names = {
        "pt": {"MM": "Muito Melhor", "M": "Melhor", "LM": "Ligeiramente Melhor", "N": "Neutro", "LP": "Ligeiramente Pior", "P": "Pior", "MP": "Muito Pior"},
        "en": {"MM": "Much Better", "M": "Better", "LM": "Slightly Better", "N": "Neutral", "LP": "Slightly Worse", "P": "Worse", "MP": "Much Worse"},
        "es": {"MM": "Mucho Mejor", "M": "Mejor", "LM": "Ligeramente Mejor", "N": "Neutro", "LP": "Ligeramente Peor", "P": "Peor", "MP": "Mucho Peor"},
        "tr": {"MM": "Çok Daha İyi", "M": "Daha İyi", "LM": "Biraz Daha İyi", "N": "Nötr", "LP": "Biraz Daha Kötü", "P": "Daha Kötü", "MP": "Çok Daha Kötü"},
    }
    return names.get(current_language(), names["pt"]).get(level_code, fallback)


st.set_page_config(page_title="RETAILL MCDA", page_icon="📊", layout="wide")


def apply_css() -> None:
    st.markdown(f"""
    <style>
    .stApp {{ background: linear-gradient(180deg, {LIGHT_BG} 0%, #FFFFFF 100%); }}
    .block-container {{ padding-top: 0.45rem; padding-bottom: 0.35rem; }}
    h1, h2, h3 {{ color: {DARK_BLUE}; letter-spacing: -0.02em; }}
    h1 {{ margin-bottom: 0.2rem; }}
    h2 {{ margin-top: 0.3rem; margin-bottom: 0.25rem; }}
    .hero {{
        background: linear-gradient(135deg, {DARK_BLUE} 0%, #1B4B6F 55%, {TEAL} 100%);
        color: white; border-radius: 20px; padding: 18px 20px;
        box-shadow: 0 14px 30px rgba(23,58,94,0.18); margin-bottom: 10px;
        position: relative; overflow: hidden;
    }}
    .hero:after {{
        content: ""; position: absolute; width: 260px; height: 260px; border-radius: 50%;
        border: 42px solid rgba(42,198,182,0.22); right: -80px; top: -80px;
    }}
    .hero-title {{ font-size: 1.9rem; font-weight: 800; margin: 0; line-height: 1.08; }}
    .hero-subtitle {{ color: #CDEFE7; font-size: 0.98rem; margin-top: 8px; max-width: 780px; }}
    .tag {{
        display: inline-block; padding: 6px 12px; border-radius: 999px;
        background: rgba(73,176,81,0.95); color: white; font-size: 0.82rem;
        font-weight: 700; margin-bottom: 16px; letter-spacing: 0.06em; text-transform: uppercase;
    }}
    .metric-card {{
        background: white; border: 1px solid rgba(23,58,94,0.08); border-left: 8px solid {GREEN};
        border-radius: 16px; padding: 18px 18px; box-shadow: 0 10px 26px rgba(23,58,94,0.08);
        min-height: 112px;
    }}
    .metric-card h4 {{ margin: 0 0 6px 0; color: {DARK_BLUE}; font-size: 0.95rem; }}
    .metric-card p {{ margin: 0; color: #466; font-size: 0.88rem; }}
    .top-logo-row {{ display: flex; flex-wrap: wrap; align-items: center; gap: 10px; margin: 0 0 12px 0; }}
    .top-logo-row img {{ max-height: 34px; width: auto; object-fit: contain; filter: drop-shadow(0 4px 8px rgba(23,58,94,0.12)); }}
    .top-logo-bar {{ display:flex; align-items:center; gap:12px; min-height:118px; padding: 18px 0 16px 0; margin: 4px 0 12px 0; flex-wrap:wrap; }}
    .top-logo-bar img {{ max-height:46px; max-width:132px; width:auto; object-fit:contain; display:block; filter: drop-shadow(0 4px 8px rgba(23,58,94,0.12)); }}
    .nav-compact {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 8px; margin: 0 0 10px 0; }}
    .nav-chip {{
        display: inline-flex; align-items: center; justify-content: center;
        min-height: 36px; padding: 6px 12px; border-radius: 999px;
        font-size: 0.85rem; line-height: 1.05; font-weight: 800;
        text-align: center; text-decoration: none; white-space: nowrap;
        color: #FFFFFF !important;
        border: 1px solid {DARK_BLUE};
        background: linear-gradient(180deg, #19506B 0%, #15465D 100%);
        box-shadow: 0 6px 12px rgba(23,58,94,0.12);
    }}
    .nav-chip.active {{
        background: linear-gradient(135deg, {GREEN} 0%, #2B8A3E 100%);
        border-color: {GREEN};
        box-shadow: 0 8px 14px rgba(73,176,81,0.18);
    }}
    div[data-testid="stButton"] button[kind="secondary"] {{
        color: #FFFFFF !important;
        border: 1px solid {DARK_BLUE};
        background: linear-gradient(180deg, #19506B 0%, #15465D 100%);
        box-shadow: 0 6px 12px rgba(23,58,94,0.12);
        font-weight: 800;
    }}
    div[data-testid="stButton"] button[kind="primary"] {{
        color: #FFFFFF !important;
        border: 1px solid {GREEN};
        background: linear-gradient(135deg, {GREEN} 0%, #2B8A3E 100%);
        box-shadow: 0 8px 14px rgba(73,176,81,0.18);
        font-weight: 800;
    }}
    @media (max-width: 760px) {{
        .nav-compact {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
        .nav-chip {{ white-space: normal; }}
    }}
    .step-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin: 10px 0 8px 0; }}
    .step-card {{
        background: linear-gradient(180deg, #19506B 0%, #15465D 100%); color: white;
        border-radius: 16px; padding: 12px 12px; min-height: 110px;
        border: 1px solid rgba(42,198,182,0.22); box-shadow: 0 10px 20px rgba(23,58,94,0.14);
    }}
    .step-number {{
        width: 40px; height: 40px; border-radius: 999px; background: {BRIGHT_TEAL};
        display: flex; align-items: center; justify-content: center; font-weight: 800;
        color: {DARK_BLUE}; margin-bottom: 10px; font-size: 0.92rem;
    }}
    .step-card.active {{ background: linear-gradient(135deg, {GREEN} 0%, #2B8A3E 100%); }}
    .winner-box {{
        background: linear-gradient(135deg, #76C879 0%, {GREEN} 100%); color: white;
        padding: 12px 14px; border-radius: 14px; font-weight: 800;
        box-shadow: 0 8px 18px rgba(73,176,81,0.18); margin-top: 6px;
    }}
    </style>
    """, unsafe_allow_html=True)


def init_state() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    st.session_state.setdefault("language", "pt")
    st.session_state.setdefault("active_page", 0)
    st.session_state.setdefault("analysis_revision", 0)
    st.session_state.setdefault("criteria_revision", 0)
    st.session_state.setdefault("active_dialog", None)
    st.session_state.setdefault("post_save_action", None)
    st.session_state.setdefault("new_analysis_name", "Nova análise")
    st.session_state.setdefault("save_analysis_filename", "")
    sample_path = DATA_DIR / "fornecedores_fruta_alpha.json"
    if not sample_path.exists():
        save_analysis(create_fruit_supplier_analysis(), sample_path)
    if "analysis" not in st.session_state:
        st.session_state.analysis = load_analysis(sample_path)
        st.session_state.current_path = str(sample_path)


def rerun_save(path: str | Path | None = None):
    saved_path = save_analysis(st.session_state.analysis, path or st.session_state.get("current_path"))
    st.session_state.current_path = str(saved_path)
    st.toast(t("saved_to", name=saved_path.name))


def safe_json_path(filename: str) -> Path:
    name = Path(filename or "").name.strip()
    if not name.lower().endswith(".json"):
        name = f"{name}.json"
    safe_name = "".join(char if char.isalnum() or char in "-_ ." else "_" for char in name).strip()
    if not safe_name or safe_name == ".json":
        safe_name = "analise.json"
    return DATA_DIR / safe_name


def mark_analysis_replaced() -> None:
    st.session_state.analysis_revision = int(st.session_state.get("analysis_revision", 0)) + 1
    st.session_state.criteria_revision = int(st.session_state.get("criteria_revision", 0)) + 1


def set_empty_analysis(name: str) -> None:
    analysis_name = (name or "").strip() or "Nova análise"
    st.session_state.analysis = Analysis(name=analysis_name)
    st.session_state.current_path = str(safe_json_path(f"{analysis_name}.json"))
    mark_analysis_replaced()


def open_save_dialog(post_save_action: str | None = None) -> None:
    current_name = Path(st.session_state.get("current_path", "")).name
    st.session_state.save_analysis_filename = current_name or f"{st.session_state.analysis.name}.json"
    st.session_state.post_save_action = post_save_action
    st.session_state.active_dialog = "save"


@st.dialog("MCDA")
def confirm_new_analysis_dialog() -> None:
    st.subheader(t("save_analysis"))
    st.write(t("confirm_new_body"))
    c1, c2, c3 = st.columns(3)
    if c1.button(t("yes"), use_container_width=True):
        open_save_dialog("new_name")
        st.rerun()
    if c2.button(t("no"), use_container_width=True):
        st.session_state.active_dialog = "new_name"
        st.rerun()
    if c3.button(t("cancel"), use_container_width=True):
        st.session_state.active_dialog = None
        st.rerun()


@st.dialog("MCDA")
def new_analysis_name_dialog() -> None:
    st.subheader(t("new_analysis_name"))
    st.text_input(t("new_analysis_name"), key="new_analysis_name")
    c1, c2 = st.columns(2)
    if c1.button(t("create"), use_container_width=True):
        set_empty_analysis(st.session_state.new_analysis_name)
        st.session_state.active_dialog = None
        st.toast(t("new_analysis_created"))
        st.rerun()
    if c2.button(t("cancel"), use_container_width=True):
        st.session_state.active_dialog = None
        st.rerun()


@st.dialog("MCDA")
def save_analysis_dialog() -> None:
    st.subheader(t("save_analysis"))
    st.text_input(t("save_filename"), key="save_analysis_filename")
    c1, c2 = st.columns(2)
    if c1.button(t("save"), use_container_width=True):
        target_path = safe_json_path(st.session_state.save_analysis_filename)
        rerun_save(target_path)
        post_save_action = st.session_state.get("post_save_action")
        st.session_state.post_save_action = None
        if post_save_action == "new_name":
            st.session_state.active_dialog = "new_name"
        elif post_save_action == "read_file":
            set_empty_analysis("Análise a carregar")
            st.session_state.active_dialog = "read_file"
        else:
            st.session_state.active_dialog = None
        st.rerun()
    if c2.button(t("cancel"), use_container_width=True):
        st.session_state.post_save_action = None
        st.session_state.active_dialog = None
        st.rerun()


@st.dialog("MCDA")
def confirm_read_analysis_dialog() -> None:
    st.subheader(t("save_analysis"))
    st.write(t("confirm_read_body"))
    c1, c2, c3 = st.columns(3)
    if c1.button(t("yes"), use_container_width=True):
        open_save_dialog("read_file")
        st.rerun()
    if c2.button(t("no"), use_container_width=True):
        set_empty_analysis("Análise a carregar")
        st.session_state.active_dialog = "read_file"
        st.rerun()
    if c3.button(t("cancel"), use_container_width=True):
        st.session_state.active_dialog = None
        st.rerun()


@st.dialog("MCDA")
def read_analysis_dialog() -> None:
    st.subheader(t("read_analysis"))
    available_files = list_analysis_files()
    selected_path = st.selectbox(
        t("select_json"),
        options=[None, *available_files],
        format_func=lambda path: "Upload JSON..." if path is None else path.name,
        key="read_analysis_existing_file",
    )
    uploaded = st.file_uploader(t("select_json"), type=["json"])
    c1, c2 = st.columns(2)
    if c1.button(t("read_analysis"), use_container_width=True):
        if selected_path is None and uploaded is None:
            st.error(t("select_json_error"))
            return
        try:
            if selected_path is not None:
                st.session_state.analysis = load_analysis(selected_path)
                st.session_state.current_path = str(selected_path)
            else:
                text_from_file = uploaded.getvalue().decode("utf-8")
                st.session_state.analysis = load_analysis_from_json_text(text_from_file)
                st.session_state.current_path = str(safe_json_path(uploaded.name))
            sync_scores_with_structure()
            mark_analysis_replaced()
            rerun_save(st.session_state.current_path)
            st.session_state.active_dialog = None
            st.toast(t("analysis_loaded"))
            st.rerun()
        except Exception as e:
            st.error(t("read_error", error=e))
    if c2.button(t("cancel"), use_container_width=True):
        st.session_state.active_dialog = None
        st.rerun()


def render_active_dialog() -> None:
    active_dialog = st.session_state.get("active_dialog")
    if active_dialog == "confirm_new":
        confirm_new_analysis_dialog()
    elif active_dialog == "new_name":
        new_analysis_name_dialog()
    elif active_dialog == "save":
        save_analysis_dialog()
    elif active_dialog == "confirm_read":
        confirm_read_analysis_dialog()
    elif active_dialog == "read_file":
        read_analysis_dialog()


def render_sidebar():
    retail_logo = Path(__file__).resolve().parent / "logos" / "RETAILL_MCDA.png"
    if retail_logo.exists():
        st.sidebar.image(str(retail_logo), width=260)
    else:
        st.sidebar.title("RETAILL MCDA")
    st.sidebar.selectbox(
        t("language"),
        options=list(LANGUAGES),
        format_func=lambda code: LANGUAGES[code],
        key="language",
    )
    st.sidebar.markdown(f"**{t('current_analysis')}:** {st.session_state.analysis.name}")
    st.sidebar.divider()
    if st.sidebar.button(t("new_analysis"), use_container_width=True):
        st.session_state.active_dialog = "confirm_new"
        st.rerun()
    if st.sidebar.button(t("read_analysis"), use_container_width=True):
        st.session_state.active_dialog = "confirm_read"
        st.rerun()
    if st.sidebar.button(t("save_analysis"), use_container_width=True):
        open_save_dialog()
        st.rerun()
    st.sidebar.divider()
    filter_options = ["Top 5", "Top 10", "Top 20", "Top 100", "Todos"]
    filter_labels = {option: option for option in filter_options}
    filter_labels["Todos"] = t("all")
    st.sidebar.selectbox(
        t("producer_filter"),
        filter_options,
        index=0,
        key="producer_display_limit",
        format_func=lambda option: filter_labels[option],
        help=t("producer_filter_help"),
    )
    st.sidebar.divider()
    st.sidebar.caption(t("complete_export"))
    render_export_buttons()


def render_top_logos() -> None:
    logo_dir = Path(__file__).resolve().parent / "logos"
    if not logo_dir.exists():
        return

    logos = sorted(
        [path for path in logo_dir.iterdir() if "retaill_mcda" not in path.stem.lower()],
        key=lambda p: p.name,
    )
    if not logos:
        return

    html_parts = ['<div class="top-logo-bar">']
    for image in logos:
        suffix = image.suffix.lower()
        if suffix in {".png"}:
            mime = "image/png"
        elif suffix in {".jpg", ".jpeg"}:
            mime = "image/jpeg"
        elif suffix == ".svg":
            mime = "image/svg+xml"
        else:
            mime = "image/*"

        encoded = base64.b64encode(image.read_bytes()).decode("utf-8")
        html_parts.append(f'<img src="data:{mime};base64,{encoded}" alt="{image.name}" />')

    html_parts.append("</div>")
    st.markdown("".join(html_parts), unsafe_allow_html=True)


def render_section_nav() -> None:
    active = int(st.query_params.get("page", st.session_state.get("active_page", 0)))
    active = min(max(active, 0), len(MENU_ITEMS) - 1)
    st.session_state.active_page = active

    st.markdown('<div class="nav-compact">', unsafe_allow_html=True)
    cols = st.columns(4)
    for idx, label in enumerate(translated_menu_items()):
        with cols[idx % 4]:
            if st.button(
                f"{idx + 1}. {label}",
                key=f"nav_{idx}",
                use_container_width=True,
                type="primary" if idx == active else "secondary",
            ):
                st.session_state.active_page = idx
                st.query_params["page"] = str(idx)
                st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


def render_hero():
    analysis = st.session_state.analysis
    st.markdown(f"""
    <div class="hero">
      <div class="tag">RETAILL MCDA</div>
      <div class="hero-title">RETAILL MCDA</div>
      <div class="hero-subtitle">{analysis.description}</div>
    </div>
    """, unsafe_allow_html=True)


def render_steps(active: int = 1):
    html = '<div class="step-grid">'
    for i, name in enumerate(translated_menu_items(), start=1):
        cls = "step-card active" if i == active else "step-card"
        html += f'<div class="{cls}"><div class="step-number">{i}</div><strong>{name}</strong></div>'
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


def alternatives_dataframe(analysis):
    columns = ["Nº", "Nome", "Custo", "Selecionada para 2.ª fase", "Descrição"]
    return pd.DataFrame(
        [
            {"Nº": a.id, "Nome": a.name, "Custo": a.cost, "Selecionada para 2.ª fase": a.selected_for_phase2, "Descrição": a.description}
            for a in analysis.alternatives
        ],
        columns=columns,
    )


def criteria_dataframe(analysis):
    columns = ["Nº", "Código", "Nome", "Fase", "Selecionado", "Peso", "Custo/Risco?", "Ordem", "Descrição"]
    return pd.DataFrame(
        [
            {"Nº": c.id, "Código": c.code, "Nome": c.name, "Fase": c.phase, "Selecionado": c.selected, "Peso": c.weight, "Custo/Risco?": c.is_cost, "Ordem": c.order, "Descrição": c.description}
            for c in analysis.criteria
        ],
        columns=columns,
    )


def scale_dataframe(analysis):
    columns = ["Nº", "Código", "Nível", "Pontuação", "Cor", "Descrição"]
    scale = scale_for_analysis(analysis)
    return pd.DataFrame(
        [
            {
                "Nº": level.id,
                "Código": level.code,
                "Nível": level.name,
                "Pontuação": level.score,
                "Cor": level.color,
                "Descrição": level.description,
            }
            for level in scale
        ],
        columns=columns,
    )


def semantic_scores_dataframe(analysis):
    rows = []
    for alt in analysis.alternatives:
        row = {"Fornecedor Nº": alt.id, "Alternativa": alt.name}
        for c in analysis.criteria:
            row[c.code] = analysis.semantic_scores.get(alt.name, {}).get(c.code, neutral_level_code(scale_for_analysis(analysis)))
        rows.append(row)
    return pd.DataFrame(rows)


def apply_alternatives_dataframe(df):
    existing_by_name = {alternative.name: alternative for alternative in st.session_state.analysis.alternatives}
    used_ids = set()
    alternatives = []
    next_id = 1
    def next_available_id():
        nonlocal next_id
        while next_id in used_ids:
            next_id += 1
        value = next_id
        used_ids.add(value)
        next_id += 1
        return value

    for _, r in df.iterrows():
        name = str(r.get("Nome", "")).strip()
        if not name:
            continue
        raw_id = r.get("Nº")
        try:
            alt_id = int(raw_id)
        except (TypeError, ValueError):
            alt_id = existing_by_name.get(name, Alternative(name=name)).id
        if alt_id <= 0 or alt_id in used_ids:
            alt_id = next_available_id()
        else:
            used_ids.add(alt_id)
        alternatives.append(
            Alternative(id=alt_id, name=name, cost=float(r.get("Custo", 0) or 0), selected_for_phase2=bool(r.get("Selecionada para 2.ª fase", False)), description=str(r.get("Descrição", "") or ""))
        )
    st.session_state.analysis.alternatives = [
        alternative for alternative in alternatives
    ]


def _clean_color(value: str) -> str:
    color = str(value or "").strip()
    if not color:
        return "#EAF4EC"
    if not color.startswith("#"):
        color = f"#{color}"
    if len(color) != 7:
        return "#EAF4EC"
    try:
        int(color[1:], 16)
    except ValueError:
        return "#EAF4EC"
    return color.upper()


def apply_scale_dataframe(df) -> None:
    levels = []
    used_codes = set()
    used_ids = set()
    next_id = 1
    for _, row in df.iterrows():
        code = str(row.get("Código", "") or "").strip().upper()
        if not code or code in used_codes:
            continue
        used_codes.add(code)
        try:
            level_id = int(row.get("Nº"))
        except (TypeError, ValueError):
            level_id = 0
        if level_id <= 0 or level_id in used_ids:
            while next_id in used_ids:
                next_id += 1
            level_id = next_id
        used_ids.add(level_id)
        levels.append(
            Level(
                id=level_id,
                code=code,
                name=str(row.get("Nível", "") or code).strip(),
                score=float(row.get("Pontuação", 0) or 0),
                color=_clean_color(row.get("Cor", "#EAF4EC")),
                description=str(row.get("Descrição", "") or ""),
            )
        )

    if not levels:
        levels = [Level(id=1, code="N", name="Neutro", score=0, color="#EAF4EC")]

    st.session_state.analysis.levels = levels
    sync_scores_with_structure()
    st.session_state.criteria_revision = int(st.session_state.get("criteria_revision", 0)) + 1


def sync_scores_with_structure(rename_map: dict[str, str] | None = None) -> None:
    analysis = st.session_state.analysis
    criterion_codes = [criterion.code for criterion in analysis.criteria]
    alternative_names = [alternative.name for alternative in analysis.alternatives]
    scale = scale_for_analysis(analysis)
    valid_levels = {scale_level.code for scale_level in scale}
    neutral_code = neutral_level_code(scale)
    rename_map = rename_map or {}
    synced_semantic_scores = {}
    synced_scores = {}

    for alternative_name in alternative_names:
        synced_semantic_scores[alternative_name] = {}
        synced_scores[alternative_name] = {}
        existing_semantic = analysis.semantic_scores.get(alternative_name, {})
        existing_scores = analysis.scores.get(alternative_name, {})
        for criterion_code in criterion_codes:
            level = str(existing_semantic.get(criterion_code, "") or "")
            old_code = rename_map.get(criterion_code)
            if level not in valid_levels and old_code:
                level = str(existing_semantic.get(old_code, "") or "")
            if level not in valid_levels:
                numeric_score = existing_scores.get(criterion_code)
                if numeric_score is None and old_code:
                    numeric_score = existing_scores.get(old_code)
                level = neutral_code if numeric_score is None else closest_level_code(float(numeric_score), scale)
            synced_semantic_scores[alternative_name][criterion_code] = level
            synced_scores[alternative_name][criterion_code] = level_score(level, scale)

    analysis.semantic_scores = synced_semantic_scores
    analysis.scores = synced_scores


def apply_criteria_dataframe(df):
    def value_or_default(value, default):
        if value is None or pd.isna(value):
            return default
        return value

    existing_criteria = {criterion.code: criterion for criterion in st.session_state.analysis.criteria}
    existing_codes_by_index = [criterion.code for criterion in st.session_state.analysis.criteria]
    active_scale = scale_for_analysis(st.session_state.analysis)
    neutral_code = neutral_level_code(active_scale)
    criteria = []
    rename_map = {}
    used_codes = set()
    used_ids = set()
    next_id = 1
    for index, row in df.iterrows():
        code = str(row.get("Código", "")).strip()
        if not code or code in used_codes:
            continue
        used_codes.add(code)
        old_code = existing_codes_by_index[index] if index < len(existing_codes_by_index) else None
        existing = existing_criteria.get(code) or (existing_criteria.get(old_code) if old_code else None)
        if old_code and old_code != code and old_code in existing_criteria:
            rename_map[code] = old_code
        try:
            criterion_id = int(value_or_default(row.get("Nº"), 0))
        except (TypeError, ValueError):
            criterion_id = existing.id if existing else 0
        if criterion_id <= 0 or criterion_id in used_ids:
            while next_id in used_ids:
                next_id += 1
            criterion_id = next_id
        used_ids.add(criterion_id)
        criteria.append(
            Criterion(
                id=criterion_id,
                code=code,
                name=str(row.get("Nome", "") or code).strip(),
                phase=int(value_or_default(row.get("Fase", 1), 1)),
                selected=bool(value_or_default(row.get("Selecionado", True), True)),
                weight=float(value_or_default(row.get("Peso", 1), 0)),
                is_cost=bool(value_or_default(row.get("Custo/Risco?", False), False)),
                order=int(value_or_default(row.get("Ordem", index + 1), index + 1)),
                description=str(row.get("Descrição", "") or ""),
                neutral_level=existing.neutral_level if existing and existing.neutral_level in {level.code for level in active_scale} else neutral_code,
                best_level=existing.best_level if existing and existing.best_level in {level.code for level in active_scale} else active_scale[0].code,
                levels=active_scale,
            )
        )
    st.session_state.analysis.criteria = criteria
    sync_scores_with_structure(rename_map)
    st.session_state.criteria_revision = int(st.session_state.get("criteria_revision", 0)) + 1


def apply_semantic_scores_dataframe(df):
    semantic_scores, scores = {}, {}
    criterion_codes = [c.code for c in st.session_state.analysis.criteria]
    scale = scale_for_analysis(st.session_state.analysis)
    valid_levels = {level.code for level in scale}
    neutral_code = neutral_level_code(scale)
    for _, row in df.iterrows():
        alt_name = str(row.get("Alternativa", "")).strip()
        if not alt_name:
            continue
        semantic_scores[alt_name], scores[alt_name] = {}, {}
        for code in criterion_codes:
            level = str(row.get(code, neutral_code) or neutral_code).strip().upper()
            if level not in valid_levels:
                level = neutral_code
            semantic_scores[alt_name][code] = level
            scores[alt_name][code] = level_score(level, scale)
    st.session_state.analysis.semantic_scores = semantic_scores
    st.session_state.analysis.scores = scores


def level_color(level_code):
    for level in scale_for_analysis(st.session_state.analysis):
        if level.code == level_code:
            return level.color
    return "#EAF4EC"


def text_color_for_background(color: str) -> str:
    clean = _clean_color(color).lstrip("#")
    red, green, blue = int(clean[0:2], 16), int(clean[2:4], 16), int(clean[4:6], 16)
    luminance = (0.299 * red + 0.587 * green + 0.114 * blue) / 255
    return "white" if luminance < 0.5 else DARK_BLUE


def producer_display_limit() -> int | None:
    selected = st.session_state.get("producer_display_limit", "Top 5")
    if selected == "Top 5":
        return 5
    if selected == "Top 10":
        return 10
    if selected == "Top 20":
        return 20
    if selected == "Top 100":
        return 100
    return None


def visible_producer_names(analysis, phase: int = 1) -> list[str]:
    limit = producer_display_limit()
    if limit is None:
        return [alternative.name for alternative in analysis.alternatives]

    ranked = evaluate_phase(analysis, phase)
    if ranked.empty:
        return [alternative.name for alternative in analysis.alternatives[:limit]]
    return ranked.sort_values("Ranking")["Alternativa"].head(limit).tolist()


def filter_alternative_rows(df: pd.DataFrame, analysis, phase: int = 1) -> pd.DataFrame:
    if df.empty or "Alternativa" not in df.columns:
        return df
    names = visible_producer_names(analysis, phase)
    return df[df["Alternativa"].isin(names)].copy()


def _safe_download_name(name: str, suffix: str) -> str:
    safe_name = "".join(char if char.isalnum() or char in "-_ " else "_" for char in name).strip()
    if not safe_name:
        safe_name = "analise"
    return f"{safe_name}.{suffix}"


def evaluation_matrix_dataframe(analysis) -> pd.DataFrame:
    rows = []
    scale = scale_for_analysis(analysis)
    neutral_code = neutral_level_code(scale)
    for criterion in analysis.criteria:
        row = {"Critério Nº": criterion.id, "Critério": criterion.code, "Nome": criterion.name, "Fase": criterion.phase}
        for alternative in analysis.alternatives:
            level = analysis.semantic_scores.get(alternative.name, {}).get(criterion.code, neutral_code)
            row[f"{alternative.id}. {alternative.name}"] = f"{level} ({level_score(level, scale):+.0f})"
        rows.append(row)
    return pd.DataFrame(rows)


def numeric_scores_dataframe(analysis) -> pd.DataFrame:
    rows = []
    scale = scale_for_analysis(analysis)
    neutral_code = neutral_level_code(scale)
    for alternative in analysis.alternatives:
        row = {"Fornecedor Nº": alternative.id, "Alternativa": alternative.name}
        for criterion in analysis.criteria:
            row[f"{criterion.id}. {criterion.code}"] = analysis.scores.get(alternative.name, {}).get(criterion.code, level_score(analysis.semantic_scores.get(alternative.name, {}).get(criterion.code, neutral_code), scale))
        rows.append(row)
    return pd.DataFrame(rows)


def normalized_weights_dataframe(analysis, phase: int) -> pd.DataFrame:
    criteria = selected_criteria(analysis, phase)
    weights = normalize_weights(criteria)
    return pd.DataFrame(
        [
            {
                "Fase": phase,
                "Critério Nº": criterion.id,
                "Critério": criterion.code,
                "Nome": criterion.name,
                "Peso": criterion.weight,
                "Peso normalizado": weights.get(criterion.code, 0.0),
                "Peso normalizado (%)": weights.get(criterion.code, 0.0) * 100,
            }
            for criterion in criteria
        ]
    )


def analysis_export_tables(analysis) -> dict[str, pd.DataFrame]:
    tables: dict[str, pd.DataFrame] = {
        "Resumo": pd.DataFrame(
            [
                {"Campo": "Nome", "Valor": analysis.name},
                {"Campo": "Descrição", "Valor": analysis.description},
                {"Campo": "Contexto", "Valor": analysis.context},
                {"Campo": "Metodologia", "Valor": analysis.methodology},
                {"Campo": "Criado por", "Valor": analysis.created_by},
                {"Campo": "Notas", "Valor": analysis.notes},
                {"Campo": "Alternativas", "Valor": len(analysis.alternatives)},
                {"Campo": "Critérios", "Valor": len(analysis.criteria)},
            ]
        ),
        "Alternativas": alternatives_dataframe(analysis),
        "Critérios": criteria_dataframe(analysis),
        "Escala": pd.DataFrame([level.model_dump() for level in scale_for_analysis(analysis)]),
        "Avaliações semânticas": semantic_scores_dataframe(analysis),
        "Avaliações numéricas": numeric_scores_dataframe(analysis),
        "Matriz avaliação": evaluation_matrix_dataframe(analysis),
    }

    for phase in (1, 2):
        tables[f"Pesos fase {phase}"] = normalized_weights_dataframe(analysis, phase)
        result = evaluate_phase(analysis, phase)
        if not result.empty:
            tables[f"Resultados fase {phase}"] = result
            tables[f"Contribuições fase {phase}"] = criterion_contributions(analysis, phase)
            tables[f"Sensibilidade fase {phase}"] = sensitivity_scenarios(analysis, phase)
            tables[f"Robustez fase {phase}"] = robustness_summary(analysis, phase)
    return tables


def build_complete_csv(analysis) -> bytes:
    chunks = []
    for name, table in analysis_export_tables(analysis).items():
        chunks.append(f"### {name}\n")
        chunks.append(table.to_csv(index=False))
        chunks.append("\n")
    return "".join(chunks).encode("utf-8-sig")


def build_complete_excel(analysis) -> bytes:
    from openpyxl.styles import Font, PatternFill

    output = BytesIO()
    tables = analysis_export_tables(analysis)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, table in tables.items():
            sheet_name = name[:31]
            table.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=DARK_BLUE.replace("#", ""))
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

        matrix_sheet = writer.sheets.get("Matriz avaliação")
        if matrix_sheet is not None:
            for row in matrix_sheet.iter_rows(min_row=2, min_col=4):
                for cell in row:
                    level = str(cell.value or "N").split(" ", 1)[0]
                    color = level_color(level).replace("#", "")
                    cell.fill = PatternFill("solid", fgColor=color)
                    if level in {"MM", "M", "MP"}:
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.font = Font(bold=True, color=DARK_BLUE.replace("#", ""))
    return output.getvalue()


def build_complete_pdf(analysis) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"RETAILL MCDA - {analysis.name}", styles["Title"]), Spacer(1, 0.3 * cm)]

    for name, table in analysis_export_tables(analysis).items():
        if table.empty:
            continue
        story.append(Paragraph(name, styles["Heading2"]))
        printable = table.copy().astype(str)
        max_rows = 28 if name != "Matriz avaliação" else 18
        if len(printable) > max_rows:
            printable = printable.head(max_rows)
            truncated = True
        else:
            truncated = False
        data = [list(printable.columns)] + printable.values.tolist()
        pdf_table = Table(data, repeatRows=1)
        pdf_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(DARK_BLUE)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDE8E0")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(pdf_table)
        if truncated:
            story.append(Paragraph("Tabela truncada no PDF. A exportação Excel contém todos os registos.", styles["Italic"]))
        story.append(PageBreak())

    if story and isinstance(story[-1], PageBreak):
        story.pop()
    doc.build(story)
    return output.getvalue()


def render_export_buttons() -> None:
    analysis = st.session_state.analysis
    st.sidebar.download_button(
        t("export_csv"),
        data=build_complete_csv(analysis),
        file_name=_safe_download_name(analysis.name, "csv"),
        mime="text/csv",
        use_container_width=True,
    )
    st.sidebar.download_button(
        t("export_excel"),
        data=build_complete_excel(analysis),
        file_name=_safe_download_name(analysis.name, "xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.sidebar.download_button(
        t("export_pdf"),
        data=_safe_build_pdf(analysis),
        file_name=_safe_download_name(analysis.name, "pdf"),
        mime="application/pdf",
        use_container_width=True,
    )


def _safe_build_pdf(analysis) -> bytes:
    try:
        return build_complete_pdf(analysis)
    except ImportError as exc:
        st.sidebar.warning(f"PDF indisponível: {exc}")
        return b""


def page_dashboard():
    analysis = st.session_state.analysis
    st.header(f"1. {t('nav_dashboard')}")
    render_hero()
    c1, c2, c3 = st.columns(3)
    c1.markdown(f'<div class="metric-card"><h4>{t("problem")}</h4><p>{analysis.context}</p></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card"><h4>{t("decision_conflicts")}</h4><p>{t("decision_conflicts_text")}</p></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card"><h4>{t("scale")}</h4><p>{t("scale_text")}</p></div>', unsafe_allow_html=True)
    st.caption(t("quick_nav"))
    st.subheader(t("steps_title"))
    render_steps()
    results = evaluate_phase(analysis, 1)
    if not results.empty:
        winner, value = results.iloc[0]["Alternativa"], results.iloc[0]["Valor Global"]
        st.markdown(f'<div class="winner-box">{t("recommended_supplier")}: {winner} — {t("global_value")} {value:.1f}</div>', unsafe_allow_html=True)
        visible_results = filter_alternative_rows(results, analysis, 1)
        fig = px.bar(visible_results.sort_values("Valor Global"), x="Valor Global", y="Alternativa", orientation="h", text="Valor Global", title=t("final_ranking"), color_discrete_sequence=[GREEN])
        fig.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_color=DARK_BLUE, title_font_color=DARK_BLUE, xaxis_title=t("global_value_axis"), yaxis_title=t("supplier"), height=360)
        fig.update_traces(texttemplate="%{text:.1f}", textposition="outside", marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)


def page_analysis():
    a = st.session_state.analysis
    revision = st.session_state.get("analysis_revision", 0)
    st.header(f"2. {t('nav_analysis')}")
    col1, col2 = st.columns([2, 1])
    with col1:
        a.name = st.text_input(t("analysis_name"), a.name, key=f"analysis_name_{revision}")
        a.description = st.text_area(t("description"), a.description, height=90, key=f"analysis_description_{revision}")
        a.context = st.text_area(t("context"), a.context, height=120, key=f"analysis_context_{revision}")
        a.notes = st.text_area(t("notes"), a.notes, height=100, key=f"analysis_notes_{revision}")
    with col2:
        methodology_options = ["1fase", "2fases", "1e2fases"]
        a.methodology = st.selectbox(t("methodology"), methodology_options, index=methodology_options.index(a.methodology), key=f"analysis_methodology_{revision}")
        a.created_by = st.text_input(t("owner"), a.created_by, key=f"analysis_created_by_{revision}")
        st.metric(t("alternatives"), len(a.alternatives)); st.metric(t("criteria"), len(a.criteria))
    if st.button(t("save_analysis_changes")):
        rerun_save()


def page_alternatives():
    st.header(f"3. {t('nav_suppliers')}")
    df = alternatives_dataframe(st.session_state.analysis)
    edited = st.data_editor(
        df,
        key=f"alternatives_editor_{st.session_state.get('analysis_revision', 0)}",
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Nº": st.column_config.NumberColumn("Nº", min_value=1, step=1),
            "Nome": st.column_config.TextColumn(t("name")),
            "Custo": st.column_config.NumberColumn(t("cost"), min_value=0.0, step=1000.0, format="€ %.0f"),
            "Selecionada para 2.ª fase": st.column_config.CheckboxColumn(t("selected_phase2")),
            "Descrição": st.column_config.TextColumn(t("description")),
        },
    )
    if st.button(t("apply_suppliers")):
        apply_alternatives_dataframe(edited)
        sync_scores_with_structure()
        rerun_save(); st.rerun()


def page_criteria():
    st.header(f"4. {t('nav_criteria')}")
    st.caption(t("criteria_help"))
    edited = st.data_editor(
        criteria_dataframe(st.session_state.analysis),
        key=f"criteria_editor_{st.session_state.get('criteria_revision', 0)}",
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Nº": st.column_config.NumberColumn("Nº", min_value=1, step=1),
            "Código": st.column_config.TextColumn(t("code"), help=t("unique_criterion_help")),
            "Nome": st.column_config.TextColumn(t("name")),
            "Fase": st.column_config.SelectboxColumn(t("phase"), options=[1, 2], default=1),
            "Selecionado": st.column_config.CheckboxColumn(t("selected"), default=True),
            "Peso": st.column_config.NumberColumn(t("weight"), min_value=0.0, step=1.0, default=1.0),
            "Custo/Risco?": st.column_config.CheckboxColumn(t("cost_risk"), default=False),
            "Ordem": st.column_config.NumberColumn(t("order"), min_value=0, step=1, default=0),
            "Descrição": st.column_config.TextColumn(t("description")),
        },
    )
    if st.button(t("apply_criteria")):
        apply_criteria_dataframe(edited); rerun_save(); st.rerun()
    criteria = selected_criteria(st.session_state.analysis, 1)
    weights = normalize_weights(criteria)
    if criteria:
        weights_df = pd.DataFrame([{"Nº": c.id, t("criterion"): c.code, t("name"): c.name, t("original_weight"): c.weight, t("normalized_weight_pct"): round(weights[c.code] * 100, 1)} for c in criteria])
        st.dataframe(weights_df, use_container_width=True)
        fig = px.bar(weights_df, x=t("normalized_weight_pct"), y=t("criterion"), orientation="h", text=t("normalized_weight_pct"), hover_data=[t("name")], title=t("relative_importance"), color_discrete_sequence=[GREEN])
        fig.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_color=DARK_BLUE, height=360)
        fig.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
        st.plotly_chart(fig, use_container_width=True)


def page_levels():
    st.header(f"5. {t('nav_levels')}")
    st.caption("Edite, adicione ou remova níveis. Os códigos são usados nas avaliações; ao aplicar, avaliações incompatíveis são remapeadas para o nível neutro ou para a pontuação mais próxima.")
    edited = st.data_editor(
        scale_dataframe(st.session_state.analysis),
        key=f"scale_editor_{st.session_state.get('criteria_revision', 0)}",
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Nº": st.column_config.NumberColumn("Nº", min_value=1, step=1),
            "Código": st.column_config.TextColumn(t("code"), help="Código único do nível, por exemplo MM, N ou P."),
            "Nível": st.column_config.TextColumn(t("level")),
            "Pontuação": st.column_config.NumberColumn(t("score"), step=1.0, default=0.0),
            "Cor": st.column_config.TextColumn("Cor", help="Cor hexadecimal, por exemplo #49B051."),
            "Descrição": st.column_config.TextColumn(t("description")),
        },
    )
    if st.button("Aplicar escala"):
        apply_scale_dataframe(edited)
        rerun_save()
        st.rerun()

    level_df = scale_dataframe(st.session_state.analysis)
    if not level_df.empty:
        fig = px.bar(level_df.sort_values("Pontuação"), x="Pontuação", y="Código", orientation="h", text="Nível", title=t("semantic_scale_title"), color="Código", color_discrete_map={level.code: level.color for level in scale_for_analysis(st.session_state.analysis)})
        fig.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_color=DARK_BLUE, height=360, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)


def page_scores():
    st.header(f"6. {t('nav_scores')}")
    df = semantic_scores_dataframe(st.session_state.analysis)
    scale = scale_for_analysis(st.session_state.analysis)
    level_options = [level.code for level in scale]
    edited = st.data_editor(
        df,
        key=f"scores_editor_{st.session_state.get('criteria_revision', 0)}",
        use_container_width=True,
        column_config={
            **{c.code: st.column_config.SelectboxColumn(f"{c.id}. {c.code} — {c.name}", options=level_options, help=c.description) for c in st.session_state.analysis.criteria},
            "Fornecedor Nº": st.column_config.NumberColumn("Fornecedor Nº"),
            "Alternativa": st.column_config.TextColumn(t("alternatives")),
        },
        disabled=["Fornecedor Nº", "Alternativa"],
    )
    if st.button(t("apply_scores")):
        apply_semantic_scores_dataframe(edited); rerun_save(); st.rerun()
    st.subheader(t("colored_matrix"))
    st.caption(t("matrix_caption"))
    visible_names = visible_producer_names(st.session_state.analysis, 1)
    alternatives_by_name = {alternative.name: alternative for alternative in st.session_state.analysis.alternatives}
    visible_alternatives = [alternatives_by_name[name] for name in visible_names if name in alternatives_by_name]
    visible_criteria = selected_criteria(st.session_state.analysis, 1)
    html = f'<table style="width:100%; border-collapse: collapse; background:white; box-shadow:0 8px 22px rgba(23,58,94,0.08);"><tr style="background:#173A5E;color:white;"><th style="padding:12px;text-align:left;">{t("producer")}</th>'
    for criterion in visible_criteria:
        html += f'<th style="padding:12px;text-align:center;">{criterion.id}. {criterion.code}<br><span style="font-size:0.75rem;font-weight:500;">{criterion.name}</span></th>'
    html += '</tr>'
    for alternative in visible_alternatives:
        html += f'<tr><td style="padding:12px;border-bottom:1px solid #E8EFEA;"><strong>{alternative.id}. {alternative.name}</strong></td>'
        for criterion in visible_criteria:
            level = st.session_state.analysis.semantic_scores.get(alternative.name, {}).get(criterion.code, neutral_level_code(scale))
            color = level_color(level)
            text_color = text_color_for_background(color)
            html += f'<td style="padding:12px;text-align:center;border-bottom:1px solid #E8EFEA;background:{color};color:{text_color};font-weight:800;">{level}<br><span style="font-size:0.75rem;">({level_score(level, scale):+.0f})</span></td>'
        html += '</tr>'
    html += '</table>'
    st.markdown(html, unsafe_allow_html=True)


def page_results():
    st.header(f"7. {t('nav_results')}")
    results = evaluate_all(st.session_state.analysis)
    if not results:
        st.info(t("no_results")); return
    for phase, df in results.items():
        st.subheader(f"{t('phase')} {phase}")
        st.markdown(f'<div class="metric-card"><h4>{t("formula")}</h4><p>{t("formula_text")}</p></div>', unsafe_allow_html=True)
        st.dataframe(df, use_container_width=True)
        if not df.empty:
            winner, value = df.iloc[0]["Alternativa"], df.iloc[0]["Valor Global"]
            st.markdown(f'<div class="winner-box">{t("recommended_supplier")}: {winner} — {t("global_value")} {value:.1f}</div>', unsafe_allow_html=True)
            visible_df = filter_alternative_rows(df, st.session_state.analysis, phase)
            fig = px.bar(visible_df[["Alternativa", "Valor Global"]].sort_values("Valor Global"), x="Valor Global", y="Alternativa", orientation="h", text="Valor Global", title=f"{t('final_ranking')} — {t('phase')} {phase}", color_discrete_sequence=[GREEN])
            fig.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_color=DARK_BLUE, title_font_color=DARK_BLUE, xaxis_title=t("global_value_axis"), yaxis_title=t("supplier"), height=420)
            fig.update_traces(texttemplate="%{text:.1f}", textposition="outside")
            st.plotly_chart(fig, use_container_width=True)
            contrib = criterion_contributions(st.session_state.analysis, phase)
            visible_contrib = filter_alternative_rows(contrib, st.session_state.analysis, phase)
            fig2 = px.bar(visible_contrib, x="Alternativa", y="Contribuição", color="Critério", title=t("criteria_contribution"))
            fig2.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_color=DARK_BLUE, height=420)
            st.plotly_chart(fig2, use_container_width=True)
        st.download_button(t("download_phase_csv", phase=phase), data=df.to_csv(index=False).encode("utf-8"), file_name=f"resultados_fase_{phase}.csv", mime="text/csv")


def page_sensitivity():
    st.header(f"8. {t('nav_sensitivity')}")
    phase = st.selectbox(t("phase"), [1, 2], index=0, key=f"sensitivity_phase_{st.session_state.get('analysis_revision', 0)}")
    sens_df = sensitivity_scenarios(st.session_state.analysis, phase)
    if sens_df.empty:
        st.info(t("no_sensitivity")); return
    st.subheader(t("scenarios"))
    st.dataframe(sens_df, use_container_width=True)
    visible_sens_df = filter_alternative_rows(sens_df, st.session_state.analysis, phase)
    fig = px.line(visible_sens_df, x="Cenário", y="Valor Global", color="Alternativa", markers=True, title=t("ranking_stability"))
    fig.update_layout(plot_bgcolor="white", paper_bgcolor="white", font_color=DARK_BLUE, height=430)
    st.plotly_chart(fig, use_container_width=True)
    robust = robustness_summary(st.session_state.analysis, phase)
    st.subheader(t("robustness_summary"))
    st.dataframe(robust, use_container_width=True)

    if not robust.empty and bool(robust["Mantém vencedor base?"].all()):
        base_winner = robust.iloc[0]["Vencedor"]
        st.markdown(f'<div class="winner-box">{t("robust_conclusion", winner=base_winner)}</div>', unsafe_allow_html=True)
    else:
        st.warning(t("sensitive_ranking"))


def page_json():
    st.header("9. JSON")
    uploaded = st.file_uploader("Importar JSON", type=["json"])
    if uploaded is not None and st.button("Carregar ficheiro JSON"):
        try:
            text_from_file = uploaded.getvalue().decode("utf-8")
            st.session_state.analysis = load_analysis_from_json_text(text_from_file)
            sync_scores_with_structure()
            mark_analysis_replaced()
            rerun_save()
            st.success("JSON importado com sucesso.")
            st.rerun()
        except Exception as e:
            st.error(f"Erro ao importar JSON: {e}")
    text = st.text_area("JSON da análise", analysis_to_json(st.session_state.analysis), height=500)
    c1, c2 = st.columns(2)
    if c1.button("Validar e carregar JSON"):
        try:
            st.session_state.analysis = load_analysis_from_json_text(text); sync_scores_with_structure(); mark_analysis_replaced(); rerun_save(); st.success("JSON carregado com sucesso."); st.rerun()
        except Exception as e:
            st.error(f"Erro ao validar JSON: {e}")
    c2.download_button("Descarregar JSON atual", data=analysis_to_json(st.session_state.analysis), file_name=f"{st.session_state.analysis.name}.json", mime="application/json")


def main():
    init_state(); apply_css(); render_sidebar(); render_active_dialog()

    selected_page = int(st.query_params.get("page", st.session_state.get("active_page", 0)))
    selected_page = min(max(selected_page, 0), len(MENU_ITEMS) - 1)
    st.session_state.active_page = selected_page

    render_top_logos()
    render_section_nav()

    pages = [
        page_dashboard,
        page_analysis,
        page_alternatives,
        page_criteria,
        page_levels,
        page_scores,
        page_results,
        page_sensitivity,
    ]

    pages[selected_page]()


if __name__ == "__main__":
    main()
